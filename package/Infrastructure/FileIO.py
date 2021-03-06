# -*- coding: utf-8 -*-
# 檔案路徑相關操作
import os
import shutil  # file io
from openpyxl import Workbook
from openpyxl import load_workbook


class FileIO:
    """自定義檔案物件"""

    def __init__(self, fileName, dateStr):
        """"
        fileName:檔案名稱
        dateStr:日期字串
        """

        SaveDirectory = os.getcwd()  # 印出目前工作目錄
        SaveAs = os.path.join(SaveDirectory, 'daily',
                              fileName + '.xlsx')  # 組合路徑，自動加上兩條斜線 "\\"
        backupSaveAs = os.path.join(SaveDirectory, 'backup',
                                    fileName + dateStr+'.xlsx')  # 備份組合路徑，自動加上兩條斜線 "\\"

        # csv file exist or not
        # os.path.isfile('test.txt') #如果不存在就返回False
        # os.path.exists(directory) #如果目錄不存在就返回False
        self.SaveAs = SaveAs  # 檔案路徑
        self.BackupSaveAs = backupSaveAs  # 備份檔案路徑
        self.XlsExist = os.path.isfile(SaveAs)  # 檔案是否存在
        
    def CopyFile(self):
        """備份檔案"""
        shutil.copy(self.SaveAs, self.BackupSaveAs)
        # os.rename(SaveAs, backupSaveAs)#檔案搬走 (剪下貼上概念)

    def GetWorkBook(self):
        """取得或新增 excel"""
        if self.XlsExist:
            # 先備份檔案
            self.CopyFile()
            return load_workbook(filename=self.SaveAs)
        else:
            return Workbook()